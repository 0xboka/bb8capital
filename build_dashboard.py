import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from collections import defaultdict

# ─── COLOURS ────────────────────────────────────────────────────────────────
C_BG       = "0D1B2A"   # very dark navy
C_MED      = "14253A"   # medium dark
C_HDR      = "0F3460"   # header blue
C_ROW_ALT  = "111F33"   # alt row
C_WHITE    = "FFFFFF"
C_CYAN     = "00D4E8"
C_GREEN    = "4CAF50"
C_ORANGE   = "FF6B35"
C_YELLOW   = "FFC107"
C_RED      = "FF4D4D"
C_LIGHT    = "B0BEC5"
C_ACCENT   = "3D8BCD"

# ─── HELPER STYLE FUNCTIONS ─────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill(fill_type='solid', fgColor=hex_color)

def font(color=C_WHITE, bold=False, size=10, italic=False):
    return Font(color=color, bold=bold, size=size, name='Arial', italic=italic)

def align(h='center', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border(sides='bottom', color="1A3A5C"):
    s = Side(style='thin', color=color)
    kw = {k: s for k in (['bottom'] if sides == 'bottom'
                          else ['left','right','top','bottom'])}
    return Border(**kw)

def style_cell(cell, bg=C_BG, fg=C_WHITE, bold=False, size=10,
               h='left', v='center', wrap=False, italic=False):
    cell.fill = fill(bg)
    cell.font = font(fg, bold, size, italic)
    cell.alignment = align(h, v, wrap)

def money(v):
    try:
        return f"${float(v):,.0f}"
    except Exception:
        return str(v)

def pct(v):
    try:
        return f"{float(v)*100:.2f}%"
    except Exception:
        return str(v)

# ─── LOAD SOURCE DATA ────────────────────────────────────────────────────────
SRC = "bb8_25_dashboard_v2.xlsx"
xl  = pd.ExcelFile(SRC)
MONTHS = ['2412','2501','2502','2503','2504','2505','2506','2507','2508']
MONTH_LABELS = {
    '2412':'Dec 24','2501':'Jan 25','2502':'Feb 25','2503':'Mar 25',
    '2504':'Apr 25','2505':'May 25','2506':'Jun 25','2507':'Jul 25','2508':'Aug 25'
}

def load_monthly(sheet):
    df = xl.parse(sheet, header=None)
    # KPI: row indices (0-based)
    r   = df.iloc[2]   # values row (Excel row 3)
    start  = r.get(1, 20_000_000)
    prev   = r.get(2, 0)
    actual = r.get(3) if pd.notna(r.get(3, np.nan)) else r.get(4, 0)

    def kpi(row_idx, val_col1=2, val_col2=None):
        rr = df.iloc[row_idx] if row_idx < len(df) else {}
        v1 = rr.get(val_col1, np.nan)
        v2 = rr.get(val_col2 if val_col2 else 3, np.nan)
        if not pd.notna(v2):
            v2 = rr.get(4, np.nan)
        return (v1 if pd.notna(v1) else None,
                v2 if pd.notna(v2) else None)

    diff_val  = actual - start if pd.notna(actual) and pd.notna(start) else None
    diff_pct  = diff_val / start if diff_val and start else None
    hf_min    = kpi(21)
    hf_global = kpi(24)
    debt      = kpi(27)
    apy       = kpi(30)
    non_inv   = kpi(33)
    non_evm   = kpi(36)
    altcoins  = kpi(39)

    # Position table starts at pandas row 46 (Excel row 47)
    tbl_header_row = 45  # 0-indexed
    # Find actual header row containing 'wallet'
    for i in range(40, 55):
        if i < len(df) and df.iloc[i, 1] == 'wallet':
            tbl_header_row = i
            break

    tbl = df.iloc[tbl_header_row+1:].copy()
    tbl.columns = range(len(tbl.columns))
    # Drop fully-empty rows
    tbl = tbl.dropna(how='all')

    return {
        'start': start, 'prev': prev, 'actual': actual,
        'diff_val': diff_val, 'diff_pct': diff_pct,
        'hf_min': hf_min, 'hf_global': hf_global,
        'debt': debt, 'apy': apy,
        'non_inv': non_inv, 'non_evm': non_evm, 'altcoins': altcoins,
        'tbl': tbl, 'tbl_header_row': tbl_header_row,
    }

def load_wallet():
    return xl.parse('_wallet', header=None)

def load_data():
    return xl.parse('_data', header=None)

monthly = {m: load_monthly(m) for m in MONTHS}
wallet_raw = load_wallet()
data_raw   = load_data()

# ─── TOKEN → CATEGORY MAPPING from _wallet ──────────────────────────────────
token_cat = {}   # token → category
token_sub = {}   # token → under_category
w = wallet_raw
for i in range(len(w)):
    tok  = w.iloc[i, 1]
    cat  = w.iloc[i, 2]
    sub  = w.iloc[i, 3]
    if pd.notna(tok) and isinstance(tok, str) and tok not in ('tokens',):
        token_cat[tok.upper()] = cat if pd.notna(cat) else 'other'
        token_sub[tok.upper()] = sub if pd.notna(sub) else ''

# Also add mappings for derived token names
extra_map = {
    'WFRAX':'defi', 'CRVUSD':'stable','USDC':'stable','USDT':'stable',
    'USDT0':'stable','XDAI':'stable','FRAX':'stable',
    'WSTETH':'infra','CBETH':'infra','WETH':'infra',
    'MPENDLE':'defi','STHYPE':'perp','WHYPE':'perp',
    'SDBPT':'defi','SDMAV':'defi',
    'PT-SENA-25SEP2025':'defi',
}
for k,v in extra_map.items():
    if k not in token_cat:
        token_cat[k] = v

# ─── MONTHLY BALANCE HISTORY (for portfolio value chart) ────────────────────
balance_history = [(MONTH_LABELS[m], monthly[m]['actual']) for m in MONTHS]

# ─── BUILD WORKBOOK ──────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

# ════════════════════════════════════════════════════════════════════════════
#  WALLET SHEET
# ════════════════════════════════════════════════════════════════════════════
ws_wallet = wb.create_sheet("wallet")
ws_wallet.sheet_view.showGridLines = False

# Paint entire sheet dark
for row in ws_wallet.iter_rows(min_row=1, max_row=60, min_col=1, max_col=30):
    for cell in row:
        cell.fill = fill(C_BG)

# Title
ws_wallet.merge_cells('B1:V1')
t = ws_wallet['B1']
t.value = "WALLET TARGETS"
style_cell(t, C_HDR, C_WHITE, bold=True, size=14, h='center')
ws_wallet.row_dimensions[1].height = 28

# Headers row 3
w_headers = ['token','category','under category','','min upside from now',
             'target -','target +','oncoming unlocks','unlocking amount','',
             '0225_target','0225_current','comment','',
             '0625_target','0625_current','comment','',
             '0725_target','0725_current','comment','',
             '0825_target','0825_current','comment']
hdr_cols  = list('BCDEFGHIJKLMNOPQRSTUVWXY')
ws_wallet.row_dimensions[3].height = 36

# Section labels row 2
for grp, col_start, label in [
    (None,     'K', 'wallet target v8 (2/25)'),
    (None,     'O', 'wallet target v9 (6/25)'),
    (None,     'S', 'wallet target v9 (7/25)'),
    (None,     'W', 'wallet target v10 (8/25)'),
]:
    c = ws_wallet[f'{col_start}2']
    c.value = label
    style_cell(c, C_HDR, C_YELLOW, bold=True, size=9, h='center')

for i, (hdr, col) in enumerate(zip(w_headers, hdr_cols)):
    c = ws_wallet[f'{col}3']
    c.value = hdr
    style_cell(c, C_HDR, C_CYAN, bold=True, size=9, h='center', wrap=True)

# Add v10 (0825) target from most recent month data
# Build: token | cat | subcat | min_upside | target- | target+ | unlock | unlock_amt | ...
# plus 0225, 0625, 0725 from _wallet, plus 0825 from 2508 position data

# Get all wallet target columns from raw wallet data
w_data = wallet_raw.iloc[2:]  # skip to data rows (row index 2 = Excel row 3 = headers)

# Find column indices
#   col 1=token, 2=cat, 3=sub, 5=min_upside, 6=target-, 7=target+, 8=unlock, 9=unlock_amt
#   col 11=0225_target, 12=0225_current, 13=comment
#   col 15=0625_target, 16=0625_current, 17=comment
#   col 19=0725_target, 20=0725_current, 21=comment

# Build 0825 current from 2508 position table summed by token simplified
pos_2508 = monthly['2508']['tbl']
# col 4 = token(simplified), col 12 = value(a), col 13 = part%(a)
token_sums_2508 = defaultdict(float)
pct_sums_2508   = defaultdict(float)
for _, row_data in pos_2508.iterrows():
    tok = str(row_data.get(4, '')).strip().upper()
    val = row_data.get(12, 0)
    pct_a = row_data.get(13, 0)
    if pd.notna(val) and tok and tok not in ('NAN',''):
        token_sums_2508[tok] += float(val)
        if pd.notna(pct_a):
            pct_sums_2508[tok] += float(pct_a)

actual_2508 = monthly['2508']['actual']

data_row = 4
WALLET_ROW_START = 4  # Excel row for first data row

for i in range(3, len(wallet_raw)):
    row_w = wallet_raw.iloc[i]
    tok = row_w.get(1)
    if not pd.notna(tok) or not isinstance(tok, str):
        continue
    tok = tok.strip()
    if tok in ('tokens', ''):
        continue

    r = data_row
    ws_wallet.row_dimensions[r].height = 18
    bg = C_MED if (data_row % 2 == 0) else C_BG

    def wc(col_letter, val, fmt_fn=None, bg_override=None, fg=C_LIGHT,
           bold=False, h='center'):
        c = ws_wallet[f'{col_letter}{r}']
        c.value = val if not fmt_fn else None
        if val is not None and val != '':
            if fmt_fn:
                c.value = val
        style_cell(c, bg_override or bg, fg, bold=bold, h=h)

    # raw values from _wallet
    cat     = row_w.get(2)
    sub     = row_w.get(3)
    min_up  = row_w.get(5)
    tgt_min = row_w.get(6)
    tgt_max = row_w.get(7)
    unlock  = row_w.get(8)
    unlock_amt = row_w.get(9)
    t0225   = row_w.get(11)
    c0225   = row_w.get(12)
    cm0225  = row_w.get(13)
    t0625   = row_w.get(15)
    c0625   = row_w.get(16)
    cm0625  = row_w.get(17)
    t0725   = row_w.get(19)
    c0725   = row_w.get(20)
    cm0725  = row_w.get(21)

    # 0825 current from actual 2508 data
    tok_up  = tok.upper()
    c0825   = pct_sums_2508.get(tok_up, None)

    # format helpers
    def fv(v):
        if not pd.notna(v) or v in ('/', ''):
            return None
        try:
            return round(float(v), 6)
        except (TypeError, ValueError):
            return v

    vals = [
        ('B', tok,        C_YELLOW, True,  'left'),
        ('C', cat if pd.notna(cat) else None,  C_LIGHT, False, 'left'),
        ('D', sub if pd.notna(sub) else None,  C_LIGHT, False, 'left'),
        ('E', None,       C_BG, False, 'center'),
        ('F', fv(min_up), C_LIGHT, False, 'right'),
        ('G', fv(tgt_min),C_LIGHT, False, 'right'),
        ('H', fv(tgt_max),C_LIGHT, False, 'right'),
        ('I', fv(unlock) if unlock not in ('/',None) else None, C_LIGHT, False, 'center'),
        ('J', fv(unlock_amt) if unlock_amt not in ('/',None) else None, C_LIGHT, False, 'right'),
        ('K', None,       C_BG, False, 'center'),
        ('L', fv(t0225),  C_CYAN, False, 'right'),
        ('M', fv(c0225),  C_LIGHT, False, 'right'),
        ('N', str(cm0225) if pd.notna(cm0225) else '', C_LIGHT, False, 'left'),
        ('O', None,       C_BG, False, 'center'),
        ('P', fv(t0625),  C_CYAN, False, 'right'),
        ('Q', fv(c0625),  C_LIGHT, False, 'right'),
        ('R', str(cm0625) if pd.notna(cm0625) else '', C_LIGHT, False, 'left'),
        ('S', None,       C_BG, False, 'center'),
        ('T', fv(t0725),  C_CYAN, False, 'right'),
        ('U', fv(c0725),  C_LIGHT, False, 'right'),
        ('V', str(cm0725) if pd.notna(cm0725) else '', C_LIGHT, False, 'left'),
        ('W', None,       C_BG, False, 'center'),
        ('X', fv(t0725),  C_CYAN, False, 'right'),   # 0825 target ≈ same as 0725
        ('Y', round(c0825, 6) if c0825 else None, C_GREEN if c0825 else C_LIGHT, False, 'right'),
    ]

    for col_l, val, fg_c, bld, h_a in vals:
        c = ws_wallet[f'{col_l}{r}']
        c.value = val
        style_cell(c, bg, fg_c, bold=bld, h=h_a)
        if isinstance(val, float) and 'target' in col_l.lower():
            c.number_format = '0.0%'

    # Number formats for percentages
    for pcol in ('L','M','P','Q','T','U','X','Y'):
        c = ws_wallet[f'{pcol}{r}']
        if isinstance(c.value, float):
            c.number_format = '0.00%'

    data_row += 1

# Total row
tot_row = data_row
ws_wallet[f'B{tot_row}'].value = 'TOTAL'
style_cell(ws_wallet[f'B{tot_row}'], C_HDR, C_YELLOW, bold=True)
for col_l in ('L','M','P','Q','T','U','X','Y'):
    first_r = WALLET_ROW_START
    last_r  = data_row - 1
    c = ws_wallet[f'{col_l}{tot_row}']
    c.value = f'=SUM({col_l}{first_r}:{col_l}{last_r})'
    style_cell(c, C_HDR, C_YELLOW, bold=True, h='right')
    c.number_format = '0.00%'

# Column widths - wallet sheet
ws_wallet.column_dimensions['A'].width = 3
ws_wallet.column_dimensions['B'].width = 12
ws_wallet.column_dimensions['C'].width = 10
ws_wallet.column_dimensions['D'].width = 12
ws_wallet.column_dimensions['E'].width = 2
ws_wallet.column_dimensions['F'].width = 9
ws_wallet.column_dimensions['G'].width = 8
ws_wallet.column_dimensions['H'].width = 8
ws_wallet.column_dimensions['I'].width = 12
ws_wallet.column_dimensions['J'].width = 10
ws_wallet.column_dimensions['K'].width = 2
for col in ('L','M','N','O','P','Q','R','S','T','U','V','W','X','Y'):
    ws_wallet.column_dimensions[col].width = 10 if col in ('N','R','V') else 9

# ════════════════════════════════════════════════════════════════════════════
#  DATA SHEET
# ════════════════════════════════════════════════════════════════════════════
ws_data = wb.create_sheet("data")
ws_data.sheet_view.showGridLines = False
for row in ws_data.iter_rows(min_row=1, max_row=600, min_col=1, max_col=30):
    for cell in row:
        cell.fill = fill(C_BG)

# ── Section 1: Monthly balance history ───────────────────────────────────────
ws_data['B1'].value = "MONTHLY PORTFOLIO HISTORY"
style_cell(ws_data['B1'], C_HDR, C_YELLOW, bold=True, size=12, h='center')
ws_data.merge_cells('B1:F1')

hdr = ['Month','Starting Balance ($)','Previous Balance ($)','Actual Balance ($)',
       'Diff ($)','Diff (%)']
for j, h in enumerate(hdr, start=2):
    c = ws_data.cell(2, j)
    c.value = h
    style_cell(c, C_HDR, C_CYAN, bold=True, size=9, h='center', wrap=True)
    ws_data.row_dimensions[2].height = 28

HIST_ROW_START = 3
for i, m in enumerate(MONTHS):
    r   = HIST_ROW_START + i
    md  = monthly[m]
    s   = md['start']
    p   = md['prev']
    a   = md['actual']
    dv  = md['diff_val']
    dp  = md['diff_pct']
    vals = [MONTH_LABELS[m], s, p, a, dv, dp]
    bg = C_MED if i % 2 else C_BG
    for j, v in enumerate(vals, start=2):
        c = ws_data.cell(r, j)
        c.value = v
        style_cell(c, bg, C_WHITE if j == 2 else C_LIGHT, h='center')
        if j in (3, 4, 5, 6):
            c.number_format = '$#,##0'
        elif j == 7:
            c.number_format = '0.00%'

# Total row for history section
hr = HIST_ROW_START + len(MONTHS)
ws_data.cell(hr, 2).value = 'Total / Max'
style_cell(ws_data.cell(hr, 2), C_HDR, C_YELLOW, bold=True)

# ── Section 2: Token category mapping ────────────────────────────────────────
cat_row_start = HIST_ROW_START + len(MONTHS) + 3
ws_data.cell(cat_row_start - 1, 2).value = "TOKEN → CATEGORY MAPPING"
style_cell(ws_data.cell(cat_row_start - 1, 2), C_HDR, C_YELLOW, bold=True, size=11, h='center')
ws_data.merge_cells(f'B{cat_row_start-1}:D{cat_row_start-1}')

for j, h in enumerate(['Token','Category','Under Category'], start=2):
    c = ws_data.cell(cat_row_start, j)
    c.value = h
    style_cell(c, C_HDR, C_CYAN, bold=True, size=9, h='center')

for i, (tok, cat) in enumerate(sorted(token_cat.items())):
    r = cat_row_start + 1 + i
    bg = C_MED if i % 2 else C_BG
    ws_data.cell(r, 2).value = tok
    style_cell(ws_data.cell(r, 2), bg, C_YELLOW, bold=True)
    ws_data.cell(r, 3).value = cat
    style_cell(ws_data.cell(r, 3), bg, C_LIGHT)
    ws_data.cell(r, 4).value = token_sub.get(tok, '')
    style_cell(ws_data.cell(r, 4), bg, C_LIGHT)

# ── Section 3: Daily portfolio history from _data ────────────────────────────
# data_raw cols 18-22: date, amount($), diff($), volatility(%), ath_year($)
daily_section_row = cat_row_start + len(token_cat) + 3
ws_data.cell(daily_section_row - 1, 2).value = "DAILY PORTFOLIO HISTORY"
style_cell(ws_data.cell(daily_section_row - 1, 2), C_HDR, C_YELLOW, bold=True, size=11, h='center')
ws_data.merge_cells(f'B{daily_section_row-1}:G{daily_section_row-1}')

daily_hdrs = ['Date','Amount ($)','Diff ($)','Volatility (%)','ATH Year ($)','ATL Year ($)']
for j, h in enumerate(daily_hdrs, start=2):
    c = ws_data.cell(daily_section_row, j)
    c.value = h
    style_cell(c, C_HDR, C_CYAN, bold=True, size=9, h='center')

dr = daily_section_row + 1
for i in range(len(data_raw)):
    row_d = data_raw.iloc[i]
    dt  = row_d.get(18)
    amt = row_d.get(19)
    if pd.notna(dt) and pd.notna(amt) and isinstance(dt, (pd.Timestamp,)) and isinstance(amt, float):
        bg = C_MED if (dr % 2 == 0) else C_BG
        daily_vals = [dt, amt, row_d.get(20), row_d.get(21), row_d.get(22), row_d.get(23)]
        for j, v in enumerate(daily_vals, start=2):
            c = ws_data.cell(dr, j)
            if pd.notna(v):
                c.value = v if not isinstance(v, float) or j != 2 else round(v, 2)
            style_cell(c, bg, C_LIGHT, h='center')
            if j == 2:
                c.number_format = 'YYYY-MM-DD'
            elif j in (3, 4, 6):
                c.number_format = '$#,##0'
            elif j == 5:
                c.number_format = '0.00%'
        dr += 1

# Data sheet column widths
for col, w in [('A',3),('B',14),('C',16),('D',16),('E',16),('F',14),('G',14)]:
    ws_data.column_dimensions[col].width = w

# ════════════════════════════════════════════════════════════════════════════
#  MONTHLY DASHBOARD SHEETS
# ════════════════════════════════════════════════════════════════════════════
# Position table column headers
POS_COLS = ['wallet','token','token (1/2)','token (simplified)',
            'qty (prev)','price (prev)','value (prev)','part % (prev)','sell?',
            'qty (act)','price (act)','value (act)','part % (act)','apy % (act)',
            'comment','blockchain','method','used protocol']

# Chart palette colors
CHART_COLORS = [
    "0044FF","00C3FF","00E5A0","A0FF00","FFCC00",
    "FF6600","FF0066","CC00FF","6600FF","0066FF",
    "00CCFF","00FF99","99FF00","FFFF00","FF9900",
    "FF3300","FF00AA","9900FF","3300FF","0099FF"
]

KPI_TABLE_HEADER_ROW = 46  # Excel row where position table HEADERS are
KPI_DATA_START_ROW   = 47  # Excel row where position data starts
CHART_HELPER_ROW     = 300 # Start of chart helper data (safely below all position data)
DASHBOARD_END_ROW    = 42  # Last row of KPI/dashboard area

for sheet_name in MONTHS:
    md   = monthly[sheet_name]
    lbl  = MONTH_LABELS[sheet_name]
    tbl  = md['tbl']
    ws   = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    # Paint background
    for row in ws.iter_rows(min_row=1, max_row=200, min_col=1, max_col=40):
        for cell in row:
            cell.fill = fill(C_BG)

    # ── Row 1: Title bar ─────────────────────────────────────────────────────
    ws.merge_cells('B1:S1')
    title = ws['B1']
    title.value = f"BB8 PORTFOLIO DASHBOARD  ·  {lbl.upper()}"
    style_cell(title, C_HDR, C_WHITE, bold=True, size=14, h='center')
    ws.row_dimensions[1].height = 30

    # ── Row 2: Balance headers ───────────────────────────────────────────────
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 24
    for col, hdr, fg in [('B','Starting Balance', C_LIGHT),
                          ('C','Previous Balance', C_LIGHT),
                          ('D','Actual Balance',   C_CYAN)]:
        c = ws[f'{col}2']
        c.value = hdr
        style_cell(c, C_MED, fg, bold=True, size=9, h='center')

    s = md['start']
    p = md['prev']
    a = md['actual']
    for col, val, fg in [('B', s,  C_LIGHT), ('C', p, C_LIGHT), ('D', a, C_YELLOW)]:
        c = ws[f'{col}3']
        c.value = val if pd.notna(val) else ''
        style_cell(c, C_BG, fg, bold=True, size=11, h='center')
        c.number_format = '$#,##0'

    # ── KPI metrics rows 5-40 (left side, cols B-D) ──────────────────────────
    kpi_rows = [
        (5,  'différence / starting balance', md['diff_val'], md['diff_pct'], '$#,##0', '0.00%', C_GREEN if (md['diff_val'] or 0) > 0 else C_RED),
        (8,  'min (important) health factor',  md['hf_min'][1],  None, '0.00', None, C_YELLOW),
        (11, 'health factor (global)',          md['hf_global'][1], None, '0.00', None, C_CYAN),
        (14, 'debt ($)',                        md['debt'][1],   None,  '$#,##0',  None, C_RED),
        (17, 'weighted APY',                    md['apy'][0],   md['apy'][1], '0.00%', '0.00%', C_GREEN),
        (20, '$ non investi (= on wallet)',     md['non_inv'][0], md['non_inv'][1], '$#,##0', '$#,##0', C_LIGHT),
        (23, '% non EVM',                       md['non_evm'][0], md['non_evm'][1], '0.00%', '0.00%', C_LIGHT),
        (26, 'altcoins (%)',                    md['altcoins'][0], md['altcoins'][1], '0.00%', '0.00%', C_ORANGE),
    ]

    for (r, label, v1, v2, fmt1, fmt2, fg_c) in kpi_rows:
        ws.row_dimensions[r].height = 22
        ws.row_dimensions[r+1].height = 20
        # Label
        lc = ws[f'B{r}']
        lc.value = label
        style_cell(lc, C_MED, C_LIGHT, size=8, italic=True, h='left')
        ws.merge_cells(f'B{r}:D{r}')
        # Value row
        vc1 = ws[f'B{r+1}']
        vc1.value = v1 if (v1 is not None and pd.notna(v1)) else ''
        style_cell(vc1, C_BG, fg_c, bold=True, size=10, h='center')
        if v1 is not None and fmt1:
            vc1.number_format = fmt1
        if v2 is not None:
            vc2 = ws[f'C{r+1}']
            vc2.value = v2
            style_cell(vc2, C_BG, fg_c, bold=True, size=10, h='center')
            if fmt2:
                vc2.number_format = fmt2

    # ── Table separator row 43-44 ────────────────────────────────────────────
    ws.row_dimensions[43].height = 6
    ws.row_dimensions[44].height = 18
    ws.row_dimensions[45].height = 18
    ws.row_dimensions[46].height = 14

    lbl_row = ws['B44']
    lbl_row.value = f"◀ PREVIOUS MONTH  ─────────────────────────────────  ACTUAL MONTH ▶"
    style_cell(lbl_row, C_HDR, C_CYAN, bold=True, size=9, h='center', italic=True)
    ws.merge_cells('B44:S44')

    # Position table column headers - row 45
    for j, hdr in enumerate(POS_COLS, start=2):
        c = ws.cell(45, j)
        c.value = hdr
        style_cell(c, C_HDR, C_CYAN, bold=True, size=8, h='center', wrap=True)

    # ── Position data rows ───────────────────────────────────────────────────
    # col indices in tbl (0-based after dropna reset):
    # 1=wallet, 2=token, 3=tok_half, 4=tok_simple, 5=qty_p, 6=px_p, 7=val_p,
    # 8=pct_p, 9=sell, 10=qty_a, 11=px_a, 12=val_a, 13=pct_a, 14=apy_a,
    # 15=comment, 16=blockchain, 17=method, 18=used_protocol

    excel_data_start = 46
    pos_rows_written = 0
    for idx, (_, row_d) in enumerate(tbl.iterrows()):
        # Skip if looks like a sub-total row (wallet is NaN or 'grand total')
        wallet_val = row_d.get(1, '')
        if not pd.notna(wallet_val) or str(wallet_val).lower() in ('grand total', 'nan', ''):
            continue

        r = excel_data_start + pos_rows_written
        ws.row_dimensions[r].height = 14
        bg = C_MED if pos_rows_written % 2 == 0 else C_BG

        cols_map = [
            (2,  1,  C_ACCENT,  'left'),    # wallet
            (3,  2,  C_WHITE,   'center'),   # token
            (4,  3,  C_LIGHT,   'center'),   # tok 1/2
            (5,  4,  C_YELLOW,  'center'),   # tok simplified
            (6,  5,  C_LIGHT,   'right'),    # qty_p
            (7,  6,  C_LIGHT,   'right'),    # px_p
            (8,  7,  C_LIGHT,   'right'),    # val_p
            (9,  8,  C_LIGHT,   'right'),    # pct_p
            (10, 9,  C_ORANGE,  'center'),   # sell?
            (11, 10, C_LIGHT,   'right'),    # qty_a
            (12, 11, C_CYAN,    'right'),    # px_a
            (13, 12, C_YELLOW,  'right'),    # val_a
            (14, 13, C_GREEN,   'right'),    # pct_a
            (15, 14, C_CYAN,    'right'),    # apy_a
            (16, 15, C_LIGHT,   'left'),     # comment
            (17, 16, C_LIGHT,   'center'),   # blockchain
            (18, 17, C_ORANGE,  'center'),   # method
            (19, 18, C_LIGHT,   'center'),   # used_protocol
        ]

        for (excel_col, src_col, fg_c, h_a) in cols_map:
            c  = ws.cell(r, excel_col)
            v  = row_d.get(src_col)
            c.value = v if pd.notna(v) else None
            style_cell(c, bg, fg_c, h=h_a, size=8)

        # Number formats
        for col_i, fmt in [(8, '$#,##0'), (9, '0.00%'), (13, '$#,##0'),
                           (14, '0.00%'), (15, '0.00%'), (7, '$#,##0'), (12, '$#,##0')]:
            c = ws.cell(r, col_i)
            if isinstance(c.value, (int, float)):
                c.number_format = fmt

        pos_rows_written += 1

    # Grand total row
    if pos_rows_written > 0:
        tot_r = excel_data_start + pos_rows_written
        ws.row_dimensions[tot_r].height = 18
        c = ws.cell(tot_r, 2)
        c.value = 'GRAND TOTAL'
        style_cell(c, C_HDR, C_YELLOW, bold=True, size=9, h='center')
        ws.merge_cells(f'B{tot_r}:D{tot_r}')

        for col_i, src_col_letter, fmt in [
            (8, 'H', '$#,##0'), (13, 'M', '$#,##0'),
            (9, 'I', '0.00%'), (14, 'N', '0.00%'), (15, 'O', '0.00%')
        ]:
            c = ws.cell(tot_r, col_i)
            c.value = f'=SUM({src_col_letter}{excel_data_start}:{src_col_letter}{excel_data_start+pos_rows_written-1})'
            style_cell(c, C_HDR, C_YELLOW, bold=True, h='right')
            c.number_format = fmt

    # ── Chart helper data (rows 180+) ────────────────────────────────────────
    # TOKEN ALLOCATION helper
    tok_helper_row = CHART_HELPER_ROW
    ws.cell(tok_helper_row, 2).value = '__TOKEN_DATA__'
    ws.cell(tok_helper_row, 2).font  = Font(color="0D1B2A", size=6)  # invisible

    # Calculate token totals from actual position data
    tok_totals = defaultdict(float)
    for _, row_d in tbl.iterrows():
        tok_s = str(row_d.get(4, '')).strip().upper()
        val_a = row_d.get(12, 0)
        if pd.notna(val_a) and tok_s and tok_s not in ('NAN', ''):
            tok_totals[tok_s] += float(val_a) if pd.notna(val_a) else 0

    # Filter to positive values and sort descending
    tok_totals_pos = {k: v for k, v in tok_totals.items() if v > 0}
    tok_sorted = sorted(tok_totals_pos.items(), key=lambda x: -x[1])
    # Limit to top 20, group rest as "Other"
    if len(tok_sorted) > 20:
        top20 = tok_sorted[:20]
        other_val = sum(v for _, v in tok_sorted[20:])
        if other_val > 0:
            top20.append(('OTHER', other_val))
        tok_sorted = top20

    tok_label_col = 2
    tok_val_col   = 3
    tok_label_start = tok_helper_row + 1
    for i, (tok_name, tok_val) in enumerate(tok_sorted):
        r = tok_label_start + i
        ws.cell(r, tok_label_col).value = tok_name
        ws.cell(r, tok_val_col).value   = round(tok_val, 2)

    # CATEGORY ALLOCATION helper
    cat_totals = defaultdict(float)
    for tok_name, tok_val in tok_totals_pos.items():
        cat = token_cat.get(tok_name, 'other')
        cat_totals[str(cat)] += tok_val
    cat_sorted    = sorted(cat_totals.items(), key=lambda x: -x[1])
    cat_helper_row = tok_label_start + len(tok_sorted) + 2
    ws.cell(cat_helper_row, 2).value = '__CAT_DATA__'
    ws.cell(cat_helper_row, 2).font  = Font(color="0D1B2A", size=6)
    cat_label_start = cat_helper_row + 1
    for i, (cat_name, cat_val) in enumerate(cat_sorted):
        r = cat_label_start + i
        ws.cell(r, tok_label_col).value = cat_name
        ws.cell(r, tok_val_col).value   = round(cat_val, 2)

    # METHOD ALLOCATION helper
    meth_totals = defaultdict(float)
    for _, row_d in tbl.iterrows():
        meth  = str(row_d.get(17, '')).strip().lower()
        val_a = row_d.get(12, 0)
        if pd.notna(val_a) and meth and meth not in ('nan', ''):
            meth_totals[meth] += float(val_a) if pd.notna(val_a) else 0
    meth_sorted    = sorted(meth_totals.items(), key=lambda x: -x[1])
    meth_helper_row = cat_label_start + len(cat_sorted) + 2
    ws.cell(meth_helper_row, 2).value = '__METH_DATA__'
    ws.cell(meth_helper_row, 2).font  = Font(color="0D1B2A", size=6)
    meth_label_start = meth_helper_row + 1
    for i, (meth_name, meth_val) in enumerate(meth_sorted):
        r = meth_label_start + i
        ws.cell(r, tok_label_col).value = meth_name
        ws.cell(r, tok_val_col).value   = round(meth_val, 2)

    # ── CHARTS ────────────────────────────────────────────────────────────────
    def make_pie(title_str, label_row_start, label_row_end, val_col,
                 anchor_cell, width_cm=12, height_cm=12):
        chart = PieChart()
        chart.title = title_str
        chart.style = 2
        chart.dataLabels = None

        labels = Reference(ws, min_col=tok_label_col,
                           min_row=label_row_start, max_row=label_row_end)
        data   = Reference(ws, min_col=val_col,
                           min_row=label_row_start, max_row=label_row_end)
        chart.add_data(data)
        chart.set_categories(labels)
        chart.series[0].graphicalProperties.line.noFill = True

        # Color slices
        n_slices = label_row_end - label_row_start + 1
        for k in range(n_slices):
            pt = DataPoint(idx=k)
            pt.graphicalProperties.solidFill = CHART_COLORS[k % len(CHART_COLORS)]
            chart.series[0].dPt.append(pt)

        chart.width  = width_cm
        chart.height = height_cm
        ws.add_chart(chart, anchor_cell)

    # Chart 1: Token allocation pie
    if tok_sorted:
        tok_end = tok_label_start + len(tok_sorted) - 1
        make_pie(f'Portfolio Allocation / Token · {lbl}',
                 tok_label_start, tok_end, tok_val_col,
                 'F2', width_cm=12, height_cm=11)

    # Chart 2: Category allocation pie
    if cat_sorted:
        cat_end = cat_label_start + len(cat_sorted) - 1
        make_pie(f'Portfolio Allocation / Category · {lbl}',
                 cat_label_start, cat_end, tok_val_col,
                 'P2', width_cm=12, height_cm=11)

    # Chart 3: Portfolio value over time (bar chart) - same data for all months
    port_chart = BarChart()
    port_chart.type  = "col"
    port_chart.title = "Portfolio Value ($) · 2024–2025"
    port_chart.style = 2
    port_chart.grouping = "clustered"

    # Write portfolio history helper near chart helpers
    port_helper_row = meth_label_start + len(meth_sorted) + 2
    ws.cell(port_helper_row, 2).value = '__PORT_HIST__'
    ws.cell(port_helper_row, 2).font  = Font(color="0D1B2A", size=6)
    port_label_start = port_helper_row + 1
    for i, (m_lbl, bal) in enumerate(balance_history):
        r = port_label_start + i
        ws.cell(r, tok_label_col).value = m_lbl
        ws.cell(r, tok_val_col).value   = round(bal, 2) if pd.notna(bal) else 0

    port_end = port_label_start + len(balance_history) - 1
    port_data   = Reference(ws, min_col=tok_val_col,
                            min_row=port_label_start, max_row=port_end)
    port_labels = Reference(ws, min_col=tok_label_col,
                            min_row=port_label_start, max_row=port_end)
    port_chart.add_data(port_data)
    port_chart.set_categories(port_labels)
    if port_chart.series:
        port_chart.series[0].graphicalProperties.solidFill = C_CYAN
    port_chart.width  = 14
    port_chart.height = 11
    ws.add_chart(port_chart, 'AB2')

    # Chart 4: Method allocation pie
    if meth_sorted:
        meth_end = meth_label_start + len(meth_sorted) - 1
        make_pie(f'Method Allocation (%) · {lbl}',
                 meth_label_start, meth_end, tok_val_col,
                 'AB14', width_cm=14, height_cm=11)

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 14   # wallet
    ws.column_dimensions['C'].width = 10   # token
    ws.column_dimensions['D'].width = 10   # tok 1/2
    ws.column_dimensions['E'].width = 10   # tok simplified
    ws.column_dimensions['F'].width = 12   # qty_p
    ws.column_dimensions['G'].width = 10   # px_p
    ws.column_dimensions['H'].width = 12   # val_p
    ws.column_dimensions['I'].width = 8    # pct_p
    ws.column_dimensions['J'].width = 14   # sell?
    ws.column_dimensions['K'].width = 12   # qty_a
    ws.column_dimensions['L'].width = 10   # px_a
    ws.column_dimensions['M'].width = 12   # val_a
    ws.column_dimensions['N'].width = 8    # pct_a
    ws.column_dimensions['O'].width = 8    # apy_a
    ws.column_dimensions['P'].width = 18   # comment
    ws.column_dimensions['Q'].width = 14   # blockchain
    ws.column_dimensions['R'].width = 12   # method
    ws.column_dimensions['S'].width = 14   # used_protocol

# ── Freeze panes & tab colors ─────────────────────────────────────────────────
for m in MONTHS:
    ws = wb[m]
    ws.freeze_panes = f'B{KPI_TABLE_HEADER_ROW}'  # freeze above table headers
    ws.sheet_properties.tabColor = "0F3460"

wb['wallet'].sheet_properties.tabColor = "1A5F2A"
wb['data'].sheet_properties.tabColor   = "7B3F00"

# ── Sheet order ───────────────────────────────────────────────────────────────
# wallet → data → monthly sheets (already in correct order)

# ─── SAVE ────────────────────────────────────────────────────────────────────
OUT = "bb8_25_dashboard_v3.xlsx"
wb.save(OUT)
print(f"Saved: {OUT}")
print(f"Sheets: {wb.sheetnames}")
